import os
import time
import json
import pymongo
import dotenv
import openpyxl

dotenv.load_dotenv(dotenv.find_dotenv())

adapter = {
    "Mã_lớp": "MaLop",
    "Mã_lớp_kèm": "MaLopKem",
    "Mã_HP": "MaHocPhan",
    "Tên_HP": "TenHocPhan",
    "Buổi_số": "BuoiHocSo",
    "Thứ": "HocVaoThu",
    "Phòng": "PhongHoc",
    "Thời_gian": "ThoiGianHoc",
    "Tuần": "HocVaoCacTuan",
    "Loại_lớp": "LoaiLop",
    "Ghi_chú": "GhiChu",
}

prop_table = {
    "MaLop": 0,
    "MaLopKem": 0,
    "MaHocPhan": 0,
    "TenHocPhan": 0,
    "BuoiHocSo": 0,
    "HocVaoThu": 0,
    "PhongHoc": 0,
    "ThoiGianHoc": 0,
    "HocVaoCacTuan": 0,
    "LoaiLop": 0,
    "GhiChu": 0,
}

SEMESTER = os.getenv("SEMESTER")
DATABASE_NAME = os.getenv("DATABASE_NAME")
XLSX_FILEPATH = os.getenv("XLSX_PATH")
CONNECTION_STRING = os.getenv("CONNECTION_STRING")
HEAD_ROW = int(os.getenv("HEAD_ROW"))
START_DATA_ROW = int(os.getenv("START_DATA_ROW"))

client = pymongo.MongoClient(CONNECTION_STRING, serverSelectionTimeoutMS=5000)

wb = openpyxl.load_workbook(XLSX_FILEPATH)
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row

print("semester = ", SEMESTER)
print("max_column =", max_column)
print("max_row =", max_row)

# determine column name then map it to prop table
for row in ws.iter_rows(min_row=HEAD_ROW, max_row=HEAD_ROW, values_only=True):
    i = 0
    for cell in row:
        prop_name = adapter.get(cell)
        if(prop_name):
            prop_table[prop_name] = i
        i = i + 1

print("=== prop_table ===")

for key in prop_table:
    print(key, "=", prop_table[key])


class SchoolClass:
    def __init__(self, row):
        for key in prop_table:
            self.__setattr__(key, row[prop_table[key]])
        self.created = int(time.time() * 1000)
        self.semester = SEMESTER
        pass

    def __str__(self) -> str:
        result = ""
        for key in prop_table:
            result = result + key + ": " + \
                str(self.__getattribute__(key)) + ", "
        result += "semester: " + str(self.semester)
        result += "created: " + str(self.created)
        return result

    def to_dict(self):
        result = {}
        for key in prop_table:
            result[key] = self.__getattribute__(key)
        result["created"] = self.created
        result["semester"] = self.semester
        return result


count = 0
stop = True

for row in ws.iter_rows(min_row=START_DATA_ROW, max_row=START_DATA_ROW, values_only=True):
    # school_class = create_school_class(row)
    school_class = SchoolClass(row)
    print(school_class)
    value = input("is that ok ? (Y) to continue (Other) to cancel: ")
    if value == "Y":
        stop = False

if (not stop):
    count = 0
    classes = []
    for row in ws.iter_rows(min_row=START_DATA_ROW, values_only=True):
        count = count + 1
        school_class = SchoolClass(row)
        classes.append(school_class.to_dict())
    result = client[DATABASE_NAME]["school.classes"].insert_many(classes)
    print('count =', count)
    print('inserted = ', len(result.inserted_ids))
